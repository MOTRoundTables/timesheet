import os
import datetime
import openpyxl
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from google.auth.exceptions import RefreshError

# If modifying these SCOPES, delete the file token.json.
SCOPES = ['https://www.googleapis.com/auth/calendar.readonly']

def get_calendar_service():
    creds = None
    token_path = 'token.json'
    credentials_path = 'credentials.json'

    if os.path.exists(token_path):
        creds = Credentials.from_authorized_user_file(token_path, SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
            except RefreshError:
                # If refresh fails, delete the token and re-authenticate
                if os.path.exists(token_path):
                    os.remove(token_path)
                flow = InstalledAppFlow.from_client_secrets_file(credentials_path, SCOPES)
                creds = flow.run_local_server(port=0)
        else:
            flow = InstalledAppFlow.from_client_secrets_file(credentials_path, SCOPES)
            creds = flow.run_local_server(port=0)
        
        with open(token_path, 'w') as token:
            token.write(creds.to_json())

    return build('calendar', 'v3', credentials=creds)

def get_calendar_events(service, start_date, end_date):
    start_time = datetime.datetime.combine(start_date, datetime.time.min).isoformat() + 'Z'
    end_time = datetime.datetime.combine(end_date, datetime.time.max).isoformat() + 'Z'
    
    events_result = service.events().list(
        calendarId='primary',
        timeMin=start_time,
        timeMax=end_time,
        singleEvents=True,
        orderBy='startTime'
    ).execute()
    return events_result.get('items', [])

def calculate_hours(start_dt, end_dt):
    duration = end_dt - start_dt
    return round(duration.total_seconds() / 3600, 2)

def are_overlapping(event1_start, event1_end, event2_start, event2_end):
    return max(event1_start, event2_start) < min(event1_end, event2_end)

def update_excel_with_calendar_events(excel_path, events, conflict_resolution_callback):
    change_log = []
    workbook = None
    sheet = None

    if os.path.exists(excel_path):
        try:
            workbook = openpyxl.load_workbook(excel_path)
            sheet = workbook.active
        except Exception as e:
            error_message = f"Error loading existing Excel file '{excel_path}': {e}. Please check the file for corruption or ensure it's not open in another program."
            change_log.append(error_message)
            raise Exception(error_message) # Re-raise the exception to stop execution
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['שנה', 'חודש', 'יום', 'זמן התחלה', 'זמן סיום', 'שעות', 'מה'])
        change_log.append("Created new Excel file with header.")

    all_events_to_write = []
    # Read existing data from the sheet into all_events_to_write
    for row_idx in range(2, sheet.max_row + 1):
        # Check if the row is completely empty (all cells are None or empty string)
        if all(sheet.cell(row=row_idx, column=col).value is None or str(sheet.cell(row=row_idx, column=col).value).strip() == '' for col in range(1, sheet.max_column + 1)):
            continue # Skip completely empty rows
        try:
            # ### START FIX ###
            # The original code failed if the Excel file contained time objects instead of strings.
            # This updated block robustly parses existing data to handle both cases.
            year = sheet.cell(row=row_idx, column=1).value
            month = sheet.cell(row=row_idx, column=2).value
            day = sheet.cell(row=row_idx, column=3).value
            start_time_val = sheet.cell(row=row_idx, column=4).value
            end_time_val = sheet.cell(row=row_idx, column=5).value
            hours = sheet.cell(row=row_idx, column=6).value
            summary = sheet.cell(row=row_idx, column=7).value

            # Reconstruct datetime objects for conflict checking
            date = datetime.date(int(year), int(month), int(day))

            # Robustly parse time values, which might be strings or datetime.time objects from openpyxl
            start_t = start_time_val if isinstance(start_time_val, datetime.time) else datetime.datetime.strptime(str(start_time_val), '%H:%M').time()
            end_t = end_time_val if isinstance(end_time_val, datetime.time) else datetime.datetime.strptime(str(end_time_val), '%H:%M').time()

            start_dt = datetime.datetime.combine(date, start_t)
            end_dt = datetime.datetime.combine(date, end_t)

            all_events_to_write.append({
                'year': year,
                'month': month,
                'day': day,
                'start_time': start_t.strftime('%H:%M'),
                'end_time': end_t.strftime('%H:%M'),
                'hours': hours,
                'summary': summary,
                'start_dt': start_dt,
                'end_dt': end_dt
            })
            # ### END FIX ###
        except (ValueError, TypeError):
            # Skip rows with invalid data
            continue

    # Process and filter new Google Calendar events
    accepted_new_calendar_events = []
    for event in events:
        if event.get('status') == 'cancelled':
            continue

        # Check attendee response status
        attendees = event.get('attendees', [])
        user_accepted = False
        if attendees: 
            for attendee in attendees:
                if attendee.get('self') and attendee.get('responseStatus') == 'accepted':
                    user_accepted = True
                    break
        else: 
            user_accepted = True
        
        if not user_accepted:
            change_log.append(f"Skipped event '{event.get('summary', 'No Title')}' due to non-accepted response status.")
            continue

        start = event['start'].get('dateTime')
        end = event['end'].get('dateTime')
        summary = event.get('summary', 'No Title')

        # Skip all-day events (events without a dateTime)
        if not start or not end:
            change_log.append(f"Skipped all-day event '{summary}'.")
            continue

        # Parse timezone-aware datetime from Google Calendar
        event_start_dt_aware = datetime.datetime.fromisoformat(start.replace('Z', '+00:00'))
        event_end_dt_aware = datetime.datetime.fromisoformat(end.replace('Z', '+00:00'))

        # Convert to local timezone and then make them naive for comparison with Excel data
        event_start_dt = event_start_dt_aware.astimezone().replace(tzinfo=None)
        event_end_dt = event_end_dt_aware.astimezone().replace(tzinfo=None)

        date = event_start_dt.date()
        start_time_str = event_start_dt.strftime('%H:%M')
        end_time_str = event_end_dt.strftime('%H:%M')
        hours = calculate_hours(event_start_dt, event_end_dt)

        new_event_data = {
            'year': date.year,
            'month': date.month,
            'day': date.day,
            'start_time': start_time_str,
            'end_time': end_time_str,
            'hours': hours,
            'summary': summary,
            'start_dt': event_start_dt,
            'end_dt': event_end_dt
        }
        accepted_new_calendar_events.append(new_event_data)

    # Integrate new calendar events into all_events_to_write, handling conflicts
    for new_event_data in accepted_new_calendar_events:
        is_duplicate = False
        for existing_event in all_events_to_write: # Check against current data
            if (existing_event['year'] == new_event_data['year'] and
                existing_event['month'] == new_event_data['month'] and
                existing_event['day'] == new_event_data['day'] and
                existing_event['start_time'] == new_event_data['start_time'] and
                existing_event['end_time'] == new_event_data['end_time'] and
                existing_event['summary'] == new_event_data['summary']):
                is_duplicate = True
                break
        
        if is_duplicate:
            change_log.append(f"Skipped duplicate event '{new_event_data['summary']}' on {new_event_data['day']}/{new_event_data['month']}/{new_event_data['year']}.")
            continue

        conflicting_existing_events = []
        for existing_event in all_events_to_write:
            if (existing_event['year'] == new_event_data['year'] and
               existing_event['month'] == new_event_data['month'] and
               existing_event['day'] == new_event_data['day'] and
               are_overlapping(new_event_data['start_dt'], new_event_data['end_dt'],
                               existing_event['start_dt'], existing_event['end_dt'])):
                conflicting_existing_events.append(existing_event)

        if conflicting_existing_events:
            should_add_new_event = True
            events_to_remove_from_all_events = [] # Collect events to remove from the master list

            for existing_event_in_conflict in conflicting_existing_events:
                action = conflict_resolution_callback(new_event_data, existing_event_in_conflict)
                
                if 'new' in action and 'existing' in action:
                    change_log.append(f"User chose to keep both new event '{new_event_data['summary']}' and existing event '{existing_event_in_conflict['summary']}' on {new_event_data['day']}/{new_event_data['month']}/{new_event_data['year']}.")
                    # No change to all_events_to_write for existing_event_in_conflict as it's kept
                elif 'new' in action:
                    # User chose to keep the new event and discard existing.
                    events_to_remove_from_all_events.append(existing_event_in_conflict)
                    change_log.append(f"User chose to replace existing event '{existing_event_in_conflict['summary']}' with new event '{new_event_data['summary']}' on {new_event_data['day']}/{new_event_data['month']}/{new_event_data['year']}.")
                elif 'existing' in action:
                    # User chose to keep existing and not add new.
                    should_add_new_event = False
                    change_log.append(f"User chose to keep existing event '{existing_event_in_conflict['summary']}' and skip new event '{new_event_data['summary']}' on {new_event_data['day']}/{new_event_data['month']}/{new_event_data['year']}.")
                    break # If user chooses to keep existing and skip new, no need to check other conflicts for this new event.
                else: # action is () - user chose neither
                    should_add_new_event = False
                    events_to_remove_from_all_events.append(existing_event_in_conflict) # Remove existing if neither is chosen
                    change_log.append(f"User chose to skip new event '{new_event_data['summary']}' and remove existing '{existing_event_in_conflict['summary']}' due to conflict on {new_event_data['day']}/{new_event_data['month']}/{new_event_data['year']}.")
                    break # If user chooses to skip, no need to check other conflicts for this new event.

            # Remove events marked for removal from the master list
            for event_to_remove in events_to_remove_from_all_events:
                if event_to_remove in all_events_to_write:
                    all_events_to_write.remove(event_to_remove)

            if should_add_new_event:
                all_events_to_write.append(new_event_data)
        else:
            # No conflicts, just append the new event to the master list
            all_events_to_write.append(new_event_data)
            change_log.append(f"Added event '{new_event_data['summary']}' on {new_event_data['day']}/{new_event_data['month']}/{new_event_data['year']} at {new_event_data['start_time']}-{new_event_data['end_time']}. Duration: {new_event_data['hours']} hours.")

    # Clear the sheet (except header) and rewrite with the consolidated data
    for row_idx in range(sheet.max_row, 1, -1):
        sheet.delete_rows(row_idx)
 
    all_events_to_write.sort(key=lambda x: (x['year'], x['month'], x['day'], x['start_dt']))

    for event_data in all_events_to_write:
        sheet.append([event_data['year'], event_data['month'], event_data['day'],
                      event_data['start_time'], event_data['end_time'],
                      event_data['hours'], event_data['summary']])

    workbook.save(excel_path)
    return change_log